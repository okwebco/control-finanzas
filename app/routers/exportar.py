from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.database import get_db
from app.models import Cuenta, Transaccion
from app.dependencies import get_current_user

router = APIRouter(prefix="/api/exportar", tags=["exportar"])

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _estilo_encabezado(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


@router.get("/cuentas")
async def exportar_cuentas(
    perfil: Optional[str] = None,
    tipo: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    q = db.query(Cuenta)
    if perfil:
        q = q.filter(Cuenta.perfil == perfil)
    if tipo:
        q = q.filter(Cuenta.tipo == tipo)
    cuentas = q.order_by(Cuenta.fecha_vencimiento).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cuentas"
    _estilo_encabezado(ws, ["ID", "Perfil", "Tipo", "Concepto", "Detalle", "URL",
                             "Recurrencia", "Valor", "Moneda", "Vencimiento"])
    for c in cuentas:
        ws.append([
            c.id,
            "Personal" if c.perfil == "personal" else "Ok Web S.A.S.",
            "Por cobrar" if c.tipo == "cxc" else "Por pagar",
            c.concepto, c.detalle or "", c.url or "",
            c.recurrencia or "", c.valor, c.moneda,
            c.fecha_vencimiento.strftime("%Y-%m-%d"),
        ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = f"cuentas_{perfil or 'todas'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre}"},
    )


@router.get("/libro")
async def exportar_libro(
    perfil: Optional[str] = None,
    año: Optional[int] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    q = db.query(Transaccion)
    if perfil:
        q = q.filter(Transaccion.perfil == perfil)
    if año:
        q = q.filter(Transaccion.año == año)
    transacciones = q.order_by(Transaccion.fecha).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Libro Contable"
    _estilo_encabezado(ws, ["ID", "Perfil", "Fecha", "Categoría", "Descripción",
                             "Tipo", "Valor", "Saldo acumulado"])
    saldo = 0.0
    for t in transacciones:
        saldo += t.valor if t.tipo == "ingreso" else -t.valor
        ws.append([
            t.id,
            "Personal" if t.perfil == "personal" else "Ok Web S.A.S.",
            t.fecha.strftime("%Y-%m-%d"),
            t.categoria.nombre if t.categoria else "Sin categoría",
            t.descripcion,
            "Ingreso" if t.tipo == "ingreso" else "Egreso",
            t.valor, saldo,
        ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = f"libro_{perfil or 'todo'}_{año or 'todos'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre}"},
    )
